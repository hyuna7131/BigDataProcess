#!/usr/bin/python3

from openpyxl import load_workbook

wb = load_workbook(filename = 'student.xlsx')
ws = wb['Sheet1']

row_max = ws.max_row

for row in range(2, row_max+1):
	ws['G' + str(row)].value = (ws['C' + str(row)].value  * 0.3) + (ws['D' + str(row)].value *0.35) + (ws['E' + str(row)].value *0.34) + (ws['F' + str(row)].value)
	
wb.save('student.xlsx')

