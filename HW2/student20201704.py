#!/usr/bin/python3

from openpyxl import load_workbook

wb = load_workbook(filename = 'student.xlsx')
ws = wb['Sheet1']

row_max = ws.max_row

# total 구하기
for row in range(2, row_max+1):
	ws['G' + str(row)].value = (ws['C' + str(row)].value  * 0.3) + (ws['D' + str(row)].value *0.35) + (ws['E' + str(row)].value *0.34) + (ws['F' + str(row)].value)
	

# 학점 부여하기
total = []
for row in range(2, row_max+1):
	total.append(ws['G' + str(row)].value)
	
reverse = total[:]
reverse.sort(reverse = True)

dou = []
for index in range(len(reverse)):
	tmp = []
	for k in range(len(reverse)):
		if(reverse[index] == reverse[k]):
			tmp.append(reverse[k])
	dou.append(tmp)

i=0
k=0
A = []
B = []
C = []
dou_student_num = len(dou[0])

for r in range(len(total)):

	list1 = []
	for c in range(len(dou[r])):
		list1.append(dou[r][c])	
	
	rank = list1.index(list1[i]) + 1 + k 
	row = total.index(list1[i]) + 2
	
	if rank <= len(total)*0.3:
		ws['H' + str(row)].value = 'A'
		A.append(list1[i])
		if len(dou[i]) == i:
			i += 1
		else:
			i = 0
		
	elif rank <= len(total)*0.7:
		ws['H' + str(row)].value = 'B'
		B.append(list1[i])
		if len(dou[i]) == i:
			i += 1
		else:
			i = 0
				
	else:
		ws['H' + str(row)].value = 'C'
		C.append(list1[i])
		if len(dou[i]) == i:
			i += 1
		else:
			i = 0
			
	k += 1
	dou_student_num += len(dou[i])
	
j=0
m=0
n=0

for r in range(int(len(A)/2)):
	rowA = total.index(A[j]) + 2
	rankA = A.index(A[j]) + 1
	
	if A[j] != A[j+1]:
		ws['H' + str(rowA)].value = 'A+'
	j += 1
	
for r in range(int(len(B)/2)):
	rowB = total.index(B[m]) + 2
	rankB = B.index(B[j]) + 1
	
	if B[m] != B[m+1]:
		ws['H' + str(rowB)].value = 'B+'
	m += 1
	
for r in range(int(len(C)/2)):
	rowC = total.index(C[n]) + 2
	rankC = C.index(C[j]) + 1
	
	if C[n] != C[n+1]:
		ws['H' + str(rowC)].value = 'C+'
	n += 1
	
	
wb.save('student.xlsx')
